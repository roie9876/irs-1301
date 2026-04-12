"""Extract rental income data from xlsx spreadsheets.

Supports two known formats:
  Format A (2023+): Monthly breakdown per property across columns B-M,
      annual total in column with header "סה״כ", tax in last column.
  Format B (2022):  Simple summary with months count, monthly rent,
      cumulative total, and tax columns.

Common invariant: property names in column A, totals row at the bottom
(no property name in A), and the grand total / tax in that row.
"""

import re

import openpyxl


def extract_rental_excel(file_path: str) -> dict:
    """Parse a rental-income xlsx and return structured extraction data.

    Returns dict with keys matching RentalExcelExtraction field names,
    each value being a dict with ``value`` and ``confidence``.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]  # Always use first sheet

    # --- Detect layout ---------------------------------------------------
    # Find header row (contains "סה״כ" or "סה\"כ" or "סכום מצטבר")
    total_col = None
    tax_col = None
    header_row = None

    for r in range(1, min(ws.max_row + 1, 5)):  # headers within first 4 rows
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or "")
            if re.search(r'סה["\u05F4״]כ', val):
                total_col = c
                header_row = r
            if re.search(r'מס\b|אחוז|13%|10%', val):
                tax_col = c
                header_row = header_row or r

    # Fallback: if headers not found, use last two numeric columns
    if total_col is None:
        # Try second-to-last column with numeric data
        for c in range(ws.max_column, 0, -1):
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and v > 1000:
                    if tax_col is None:
                        tax_col = c
                    elif total_col is None:
                        total_col = c
                        break
            if total_col is not None:
                break

    # --- Extract property names ------------------------------------------
    data_start = (header_row or 1) + 1
    properties = []
    last_data_row = data_start

    for r in range(data_start, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name and str(name).strip():
            properties.append(str(name).strip())
            last_data_row = r

    # Totals row is the first row after data rows with no col-A text
    totals_row = None
    for r in range(last_data_row + 1, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if not a_val or not str(a_val).strip():
            # Check it actually has numeric data
            if total_col:
                v = ws.cell(row=r, column=total_col).value
                if isinstance(v, (int, float)):
                    totals_row = r
                    break

    # If no separate totals row, the last row IS the totals row
    if totals_row is None:
        totals_row = ws.max_row

    # --- Extract values --------------------------------------------------
    total_income = 0.0
    tax_amount = 0.0

    if total_col:
        v = ws.cell(row=totals_row, column=total_col).value
        if isinstance(v, (int, float)):
            total_income = float(v)

    if tax_col:
        v = ws.cell(row=totals_row, column=tax_col).value
        if isinstance(v, (int, float)):
            tax_amount = float(v)

    # Compute effective tax rate
    tax_rate = round(tax_amount / total_income * 100, 1) if total_income else 0.0

    # --- Extract tax year from title or filename -------------------------
    tax_year = _extract_year(ws, file_path)

    wb.close()

    return {
        "tax_year": {"value": tax_year, "confidence": 0.7 if tax_year else 0.0},
        "properties": {"value": ", ".join(properties), "confidence": 0.95},
        "total_annual_income": {"value": total_income, "confidence": 0.95},
        "tax_rate_pct": {"value": tax_rate, "confidence": 0.9},
        "tax_amount": {"value": tax_amount, "confidence": 0.95},
    }


def _extract_year(ws, file_path: str) -> int | None:
    """Try to extract the tax year from the spreadsheet title or filename."""
    # Check title row (usually row 1)
    for r in range(1, min(ws.max_row + 1, 3)):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or "")
            m = re.search(r'לשנת\s+(\d{4})', val)
            if m:
                return int(m.group(1))

    # Fallback: extract year from filename (basename only to avoid dir names)
    import os
    basename = os.path.basename(file_path)
    m = re.search(r'(\d{4})', basename)
    if m:
        return int(m.group(1))

    return None
